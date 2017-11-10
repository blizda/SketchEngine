import argparse
import math
import sys
from openpyxl import load_workbook

def sortDict(myDict):
    return dict(sorted(myDict.items(), key=lambda x: x[1], reverse=True))

def takeTop(myDict, N):
    return dict(sorted(myDict.items(), key=lambda x: x[1], reverse=True)[:N])

def readAllFitch(sheet, N):
    fitchList = []
    for k in range(3, 10):
        tempDict = {}
        i = 2
        while True:
            val = sheet.cell(row=i, column=1).value
            val2 = sheet.cell(row=i, column=k).value
            if val is None:
                break
            tempDict[val] = val2
            i += 1
        fitchList.append(tempDict)
    for i in range(len(fitchList)):
        fitchList[i] = takeTop(fitchList[i], N)
    allDict = {}
    for i in range(len(fitchList)):
        for it in fitchList[i]:
            if it in allDict:
                allDict[it] = allDict[it] + 1
            else:
                allDict[it] = 1
    return allDict

def makeCountFitch(sheet, countDict):
    sheet['G1'] = 'Число мер'
    numD = numDict(sheet)
    for it in countDict:
        row = numD[it]
        sheet.cell(row=row, column=10).value = countDict[it]
    return sheet

def numDict(sheet):
    numD = {}
    i = 2
    while True:
        val = sheet.cell(row=i, column=1).value
        if val is None:
            break
        numD[val] = i
        i += 1
    return numD

def readCountOfFitch(sheet):
    i = 2
    countDict = {}
    while True:
        val = sheet.cell(row=i, column=1).value
        val2 = sheet.cell(row=i, column=10).value
        if val2 is None:
            break
        countDict[val] = val2
        i += 1
    return countDict

def countRel(nDict, sredDict):
    relDict = {}
    for it in nDict:
        if it in sredDict:
            if nDict[it] > 0:
                relDict[it] = math.log2(1 + 5 / nDict[it]) * sredDict[it]
            else:
                relDict[it] = math.log2(6) * sredDict[it]
    return relDict

def colLen(sheet):
    i = 1
    while True:
        val = sheet.cell(row=i, column=1).value
        if val is None:
            break
        i += 1
    return i - 1

def sredRang(sheet):
    i = 2
    sredDict = {}
    while True:
        sredR = 0
        val = sheet.cell(row=i, column=1).value
        if val is None:
            break
        for k in range(3, 10):
            sredR += sheet.cell(row=i, column=k).value
        sredDict[val] = sredR / 7
        i += 1
    return sredDict

def optRang(sheet):
    i = 2
    optDict = {}
    retCoffList = [0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    while True:
        optRL = []
        optR = 0
        val = sheet.cell(row=i, column=1).value
        if val is None:
            break
        for k in range(3, 10):
            optRL.append(sheet.cell(row=i, column=k).value)
        optRL.sort()
        for k in range(len(optRL)):
            optR += optRL[k] * retCoffList[k]
        optDict[val] = optR / 7
        i += 1
    return optDict

def writeData(wb, sheet, sredRang, optRang, relRang, fileName):
    numD = numDict(sheet)
    sheet['K1'] = 'Средний ранг'
    sheet['L1'] = 'Оптимизированный ранг'
    sheet['M1'] = 'Нормированный ранг-100'
    for it in sredRang:
        row = numD[it]
        sheet.cell(row=row, column=11).value = sredRang[it]
        sheet.cell(row=row, column=12).value = optRang[it]
        sheet.cell(row=row, column=13).value = relRang[it]
    wb.save(fileName)

def createParser():
    parser = argparse.ArgumentParser()
    parser.add_argument('-n')
    parser.add_argument('-qv', default=100)
    parser.add_argument('-out', default='result.xlsx')
    return parser

if __name__ == "__main__":
    parser = createParser()
    namespace = parser.parse_args(sys.argv[1:])
    wb = load_workbook(filename = namespace.n)
    sheet = wb[wb.get_sheet_names()[0]]
    sheet = makeCountFitch(sheet, readAllFitch(sheet, namespace.qv))
    d = sredRang(sheet)
    k = optRang(sheet)
    m = countRel(readCountOfFitch(sheet), d)
    writeData(wb, sheet, d, k, m, namespace.out)