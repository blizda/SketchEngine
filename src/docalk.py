import argparse
import requests
import sys
from openpyxl import Workbook
from scipy.stats.stats import pearsonr
import numpy as np

def zahrovSosniHuiowDaun(wb, col, name):
    wb.cell(row=1, column=col).value = name
    i = 2
    for rec in myDict:
        wb.cell(row=i, column=col).value = "Пососи писос чмо"
        i += 1
    return wb

def sortDict(myDict):
    return dict(sorted(myDict.items(), key=lambda x: x[1], reverse=False))

def sortKeys(myDict):
    return dict(sorted(myDict.items(), key=lambda x: x[0], reverse=False))

def retradRang(myDict):
    i = 1
    rangDict = {}
    for item in myDict:
        rangDict[item] = i
        i += 1
    return rangDict

def retardPrint(myDict, wb, col, name):
    wb.cell(row=1, column=col).value = name
    i = 2
    for rec in myDict:
        wb.cell(row=i, column=col).value = myDict[rec]
        i += 1
    return wb

def printKoll(myDict, wb):
    wb.cell(row=1, column=1).value = 'Коллокант'
    i = 2
    newDict = sortKeys(myDict)
    for rec in newDict:
        wb.cell(row=i, column=1).value = rec
        i += 1
    return wb

def newRatardCorrel(dict1, dict2):
    list1 = []
    list2 = []
    for item in dict1:
        list1.append(dict1[item])
        list2.append(dict2[item])
    return list1, list2

def getRetardDicrt(listOfDict):
    newList = []
    for item in listOfDict:
        newList.append(retradRang(item))
    return newList

def getSortedList(listOfDict):
    newList = []
    for it in listOfDict:
        newList.append(sortKeys(it))
    return newList

def printTitleTable(listOfTitle, wb):
    i = 2
    for rec in listOfTitle:
        wb.cell(row=1, column=i).value = rec
        wb.cell(row=i, column=1).value = rec
        i += 1
    return wb

def writeCorrelTable(listOfRetDict, listOfRetNames, wb):
    myWb = printTitleTable(listOfRetNames, wb)
    for it1 in range(len(listOfRetDict)):
        for it2 in range(len(listOfRetDict)):
            myRetCorr1, myRetCorr2 = newRatardCorrel(listOfRetDict[it1], listOfRetDict[it2])
            myWb.cell(row=it1 + 2, column=it2 + 2).value = pearsonr(np.array(myRetCorr1), np.array(myRetCorr2))[0]
    return myWb

def makeListRankAndPrint(listOfList, listOfNames, listOfRangsName, num, wb):
    newList = getSortedList(getRetardDicrt(listOfList))
    newRet = getSortedList(listOfList)
    myiter = 0
    myWb = wb
    for i in range(len(newList)):
        myWb = retardPrint(newRet[i], myWb, num + myiter, listOfNames[i])
        myiter += 1
        myWb = retardPrint(newList[i], myWb, num + myiter, listOfRangsName[i])
        myiter += 1
    return myWb

def createParser():
    parser = argparse.ArgumentParser()
    parser.add_argument('-q')
    parser.add_argument('-n')
    parser.add_argument('-k')
    parser.add_argument('-cor', default='rutenten11_8_1G')
    parser.add_argument('-mit', default='300')
    parser.add_argument('-f', default='-3')
    parser.add_argument('-t', default='1')
    return parser

if __name__ == "__main__":
    parser = createParser()
    namespace = parser.parse_args(sys.argv[1:])
    base_url = 'https://api.sketchengine.co.uk/bonito/run.cgi'
    q = 'q' + namespace.q
    data = {
        'corpname': namespace.cor,
        'format': 'json',
        'q': q,
        'cfromw': namespace.f,
        'ctow': namespace.t,
        'cmaxitems': namespace.mit,
        'csortfn': 't',
        "cbgrfns": ["t", "m", "3", "l", "s", "d", "p"],
        'username': namespace.n,
        'api_key': namespace.k
    }
    collFreq = {}
    freq = {}
    tScore = {}
    MI = {}
    MI3 = {}
    logLikelihood = {}
    minSens = {}
    logDice = {}
    MILogF = {}
    d = requests.get(base_url + '/collx', params=data).json()
    for item in d['Items']:
        str = item['str']
        collFreq[str] = item['coll_freq']
        freq[str] = item['freq']
        tScore[str] = item['Stats'][0]['s']
        MI[str] = item['Stats'][1]['s']
        MI3[str] = item['Stats'][2]['s']
        logLikelihood[str] = item['Stats'][3]['s']
        minSens[str] = item['Stats'][4]['s']
        logDice[str] = item['Stats'][5]['s']
        MILogF[str] = item['Stats'][6]['s']
    collFreq = sortDict(collFreq)
    freq = sortDict(freq)
    tScore = sortDict(tScore)
    MI = sortDict(MI)
    MI3 = sortDict(MI3)
    logLikelihood = sortDict(logLikelihood)
    minSens = sortDict(minSens)
    logDice = sortDict(logDice)
    MILogF = sortDict(MILogF)
    chusList = []
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'Table'
    dest_filename = 'results.xlsx'
    #ws1 = printKoll(collFreq, ws1)
    listofDict = [collFreq, freq, tScore, MI, MI3, logLikelihood, minSens, logDice, MILogF]
    listOfNames = ['Частота Коллокации', 'Частота Коллоканта', 'T-Score', 'MI', 'MI3', 'Log Likelihood', 'Min Sensitivity', 'LogDice', 'MI Log F',]
    listOfNamesRang = ['Частота Коллокации Rung', 'Частота Коллоканта Rung', 'T-Score Rang', 'MI Rang', 'MI3 Rang', 'Log Likelihood Rang', 'Min Sensitivity Rang', 'LogDice Rang', 'MI Log F Rang',]
    #ws1 = makeListRankAndPrint(listofDict, listOfNames, listOfNamesRang, 2, ws1)
    #ws2 = wb.create_sheet(title="Correl")
    #ws2 = writeCorrelTable(getRetardDicrt(listofDict), listOfNamesRang, ws2)
    ws1 = zahrovSosniHuiowDaun(ws1, 1, "Сосни писос")
    wb.save(dest_filename)
